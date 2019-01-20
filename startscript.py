import openpyxl
from PlayerObject import PlayerObject
print ("Welcome to Fantasy Football CLI.")

path = 'C:/Programming Projects/Fantasy Football/12team.xlsx'
workbook = openpyxl.load_workbook(path)
sheet = workbook.active



#playerAttributes = [sheet.cell(row=6,column=i).value for i in range (1,11)]

line = []
QBArray = []

#prolly should use iter_columns here because not actually changing rows
for row in sheet.iter_rows(min_row = 6, min_col = 3, max_row = 37, max_col = 15):
    for cell in row:
        #This is the part of the loop where we need to make a new player object and
        #populate it based on The cells in the current row
        line.append(cell.value)
        #now we have an array called line that has all the values for the player in that row

    playerName = line[0]
    playerTeamandNumber = line[2]
    playerValue = line[9]
    playerScarcity = line[12]
    line = [] #Fixes problem where row wasnt iterating. If we destroy it each iteration
              #We can use the same indexing method above to get the different
    QB1 = PlayerObject(playerName, playerTeamandNumber, playerValue, playerScarcity)
    #print (QB1.name)
    #print (QB1.teamandNumber)
    #print (QB1.value)
    #print (QB1.scarcity)
    QBArray.append(QB1)
for i in QBArray:
    i.printAttr()
#FOR TOMORROW
#Fix outputs find out why scarcity is always 0







#Now that we have a list consisting of one player's attributes, we just need to unpack that
#Get rid of the garbage and prepare it for becoming member values of a PlayerObject class.

#NoneType isn't in the automatic list of types so:
NoneType = type(None)
